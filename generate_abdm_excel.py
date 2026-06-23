import json
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
import re

# ─── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_POST        = "49A845"   # green
C_GET         = "0078D7"   # blue
C_DELETE      = "C00000"   # red
C_PATCH       = "7030A0"   # purple
C_M1          = "FFF2CC"   # light yellow
C_M2          = "DDEEFF"   # light blue
C_M3          = "E2EFDA"   # light green
C_HIP         = "FCE4D6"   # light orange
C_HIU         = "EAD1DC"   # light pink
C_DEBUG       = "F4CCCC"   # light red
C_DIAG        = "E8F5E9"   # light teal
C_SECT_BG     = "2E4057"   # section header
C_ALT_ROW     = "F5F7FA"   # alternate row

METHOD_COLORS = {
    "POST":   "FF49A845",
    "GET":    "FF2196F3",
    "DELETE": "FFC62828",
    "PATCH":  "FF7B1FA2",
    "PUT":    "FFFF6F00",
}

def hfill(hex6):
    return PatternFill("solid", fgColor=hex6)

def bfill(hex6):
    return PatternFill("solid", fgColor=hex6)

def cell_font(bold=False, color="000000", size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def med_border():
    s = Side(style="medium", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── Load collection ──────────────────────────────────────────────────────────
with open(r"d:\Infer\ABDM_Postman_Collection.json", encoding="utf-8") as f:
    col = json.load(f)

# ─── Flatten all requests ─────────────────────────────────────────────────────
def flatten(items, folder=""):
    rows = []
    for item in items:
        name = item.get("name", "")
        if "item" in item:
            rows += flatten(item["item"], name)
        elif "request" in item:
            req = item["request"]
            method = req.get("method", "")
            url_obj = req.get("url", {})
            raw = url_obj.get("raw", "") if isinstance(url_obj, dict) else str(url_obj)
            # strip base-url variable
            path = raw.replace("{{BASE_URL}}", "").strip()
            desc = req.get("description", "").split("\n")[0][:120]
            body_raw = ""
            if isinstance(req.get("body"), dict) and req["body"].get("mode") == "raw":
                body_raw = req["body"].get("raw", "")[:300]
            headers = req.get("header", [])
            auth_h = next((h["value"] for h in headers if h["key"] == "Authorization"), "")
            # auth label
            if "Bearer {{EMR_JWT}}" in auth_h:
                auth = "EMR JWT"
            elif "Bearer {{GATEWAY_TOKEN}}" in auth_h:
                auth = "ABDM Gateway JWT (RS256)"
            else:
                auth = "None"
            # milestone
            fp = folder + " " + name
            if "M1" in folder or "ABHA" in folder or "abha" in folder.lower():
                milestone = "M1"
                row_bg = C_M1
            elif "M2" in folder or "Care Context" in folder or "Consent" in folder or "Link" in folder:
                milestone = "M2"
                row_bg = C_M2
            elif "M3" in folder or "Health Information" in folder:
                milestone = "M3"
                row_bg = C_M3
            elif "HIP Callbacks" in folder:
                milestone = "M2/M3 HIP"
                row_bg = C_HIP
            elif "HIU Callbacks" in folder:
                milestone = "M2/M3 HIU"
                row_bg = C_HIU
            elif "Debug" in folder:
                milestone = "Debug ⚠️"
                row_bg = C_DEBUG
            elif "Activity" in folder or "Diagnostics" in folder:
                milestone = "Ops"
                row_bg = C_DIAG
            else:
                milestone = "-"
                row_bg = "FFFFFF"
            # test scripts
            tests = []
            for ev in item.get("event", []):
                if ev.get("listen") == "test":
                    tests = ev.get("script", {}).get("exec", [])
            test_str = "\n".join(t for t in tests if not t.startswith("//"))[:300]
            rows.append({
                "folder": folder,
                "name": name,
                "method": method,
                "path": path,
                "auth": auth,
                "milestone": milestone,
                "description": desc,
                "request_body": body_raw,
                "test_scripts": test_str,
                "row_bg": row_bg,
            })
    return rows

rows = flatten(col["item"])

# ─── Environment variables ─────────────────────────────────────────────────────
env_rows = col.get("variable", [])

# ─── Build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Cover / Summary
# ══════════════════════════════════════════════════════════════════════════════
ws_cover = wb.create_sheet("📋 Cover")
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions["A"].width = 5
ws_cover.column_dimensions["B"].width = 40
ws_cover.column_dimensions["C"].width = 55

# Title block
ws_cover.merge_cells("B2:C2")
ws_cover["B2"] = "INFER EMR — ABDM HIP/HIU POSTMAN COLLECTION"
ws_cover["B2"].font = Font(bold=True, size=18, color=C_HEADER_FG, name="Calibri")
ws_cover["B2"].fill = hfill(C_HEADER_BG)
ws_cover["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws_cover.row_dimensions[2].height = 40

ws_cover.merge_cells("B3:C3")
ws_cover["B3"] = "Complete ABDM API Inventory | Generated: 2026-06-15"
ws_cover["B3"].font = Font(bold=False, size=11, color="444444")
ws_cover["B3"].alignment = Alignment(horizontal="center")
ws_cover.row_dimensions[3].height = 22

summary_data = [
    ("", ""),
    ("Base URL", "https://api.inferapp.online"),
    ("Frontend", "https://emr.inferapp.online/emr/"),
    ("ABDM Gateway (Sandbox)", "https://dev.abdm.gov.in/gateway"),
    ("ABHA Service (Sandbox)", "https://abhasbx.abdm.gov.in/abha/api/v3"),
    ("", ""),
    ("Total Endpoints", str(len(rows))),
    ("Folders / Workflows", "9"),
    ("Milestones Covered", "M1 · M2 · M3"),
    ("HIP Callbacks", "15 (v0.5 + v3)"),
    ("HIU Callbacks", "10 (v0.5 + v3)"),
    ("", ""),
    ("Collection Schema", "Postman Collection v2.1"),
    ("Authentication", "EMR JWT  |  ABDM RS256 JWT (JWKS)"),
    ("Encryption", "Curve25519 ECDH · AES-256-GCM · HKDF-SHA256"),
    ("Aadhaar (Sandbox)", "999941057058"),
]
for i, (k, v) in enumerate(summary_data, start=5):
    r = i + 1
    ws_cover[f"B{r}"] = k
    ws_cover[f"C{r}"] = v
    ws_cover[f"B{r}"].font = Font(bold=bool(k), size=10, color="1F3864" if k else "888888")
    ws_cover[f"C{r}"].font = Font(size=10, color="000000")
    ws_cover.row_dimensions[r].height = 18

ws_cover.merge_cells("B24:C24")
ws_cover["B24"] = "⚠️  SECURITY NOTICE — Remove /api/abdm/debug/* endpoints before production deployment"
ws_cover["B24"].font = Font(bold=True, size=10, color="FFFFFF")
ws_cover["B24"].fill = hfill("C00000")
ws_cover["B24"].alignment = Alignment(horizontal="center", vertical="center")
ws_cover.row_dimensions[24].height = 24

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Full API Inventory
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("🗂 API Inventory")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

COLS = [
    ("Folder / Workflow",   28),
    ("Request Name",        38),
    ("Method",               9),
    ("Path",                55),
    ("Authentication",      28),
    ("Milestone",           13),
    ("Description",         55),
    ("Request Body (sample)", 50),
    ("Test Scripts",        50),
]

# Header row 1 — big title
ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
ws["A1"] = "INFER EMR — ABDM API Inventory"
ws["A1"].font = Font(bold=True, size=13, color=C_HEADER_FG)
ws["A1"].fill = hfill(C_HEADER_BG)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Header row 2 — column labels
for ci, (label, width) in enumerate(COLS, start=1):
    cl = get_column_letter(ci)
    ws.column_dimensions[cl].width = width
    c = ws.cell(row=2, column=ci, value=label)
    c.font = Font(bold=True, size=10, color=C_HEADER_FG)
    c.fill = hfill("2E4057")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
ws.row_dimensions[2].height = 22

prev_folder = None
for ri, row in enumerate(rows, start=3):
    is_alt = (ri % 2 == 0)
    bg = row["row_bg"] if row["row_bg"] != "FFFFFF" else (C_ALT_ROW if is_alt else "FFFFFF")

    # Section separator row when folder changes
    if row["folder"] != prev_folder:
        prev_folder = row["folder"]
        sep_row = ri
        for ci in range(1, len(COLS)+1):
            c = ws.cell(row=sep_row, column=ci)
            c.fill = hfill(C_SECT_BG)
        ws.merge_cells(f"A{sep_row}:{get_column_letter(len(COLS))}{sep_row}")
        ws.cell(row=sep_row, column=1).value = f"  ▶  {row['folder']}"
        ws.cell(row=sep_row, column=1).font = Font(bold=True, size=10, color="FFFFFF")
        ws.cell(row=sep_row, column=1).alignment = Alignment(vertical="center")
        ws.row_dimensions[sep_row].height = 18
        ri += 1
        # redo data row index
        data_row = ws.max_row + 1
    else:
        data_row = ws.max_row + 1

    # Determine actual next row
    data_row = ws.max_row + 1

    vals = [
        row["folder"],
        row["name"],
        row["method"],
        row["path"],
        row["auth"],
        row["milestone"],
        row["description"],
        row["request_body"],
        row["test_scripts"],
    ]
    for ci, val in enumerate(vals, start=1):
        c = ws.cell(row=data_row, column=ci, value=val)
        c.border = thin_border()
        c.alignment = Alignment(vertical="top", wrap_text=(ci in (7,8,9)))
        c.font = Font(size=9)
        if ci == 3:  # Method cell
            mc = METHOD_COLORS.get(row["method"], "FF888888")
            c.fill = PatternFill("solid", fgColor=mc[2:])
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci in (8, 9):  # body / tests
            c.font = Font(size=8, name="Courier New", color="333333")
            c.fill = bfill("F8F8F8")
        else:
            c.fill = bfill(bg)

    ws.row_dimensions[data_row].height = 52

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Environment Variables
# ══════════════════════════════════════════════════════════════════════════════
ws_env = wb.create_sheet("🔧 Environment Variables")
ws_env.sheet_view.showGridLines = False
ws_env.freeze_panes = "A3"

ws_env.merge_cells("A1:D1")
ws_env["A1"] = "ABDM Postman Environment Variables"
ws_env["A1"].font = Font(bold=True, size=13, color=C_HEADER_FG)
ws_env["A1"].fill = hfill(C_HEADER_BG)
ws_env["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_env.row_dimensions[1].height = 30

env_cols = [("Variable Key", 30), ("Default Value", 50), ("Type", 12), ("Description / Notes", 60)]
for ci, (label, width) in enumerate(env_cols, start=1):
    ws_env.column_dimensions[get_column_letter(ci)].width = width
    c = ws_env.cell(row=2, column=ci, value=label)
    c.font = Font(bold=True, size=10, color=C_HEADER_FG)
    c.fill = hfill("2E4057")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

# Inline env data (from collection variables + extra context)
env_data = [
    ("BASE_URL",         "https://api.inferapp.online",              "default", "Infer EMR backend API base URL"),
    ("FRONTEND_URL",     "https://emr.inferapp.online/emr/",         "default", "Infer EMR frontend URL"),
    ("ABDM_GATEWAY_URL", "https://dev.abdm.gov.in/gateway",          "default", "ABDM Sandbox Gateway"),
    ("ABDM_HIECM_URL",   "https://dev.abdm.gov.in/api/hiecm",        "default", "ABDM HIECM v3"),
    ("ABHA_BASE_URL",    "https://abhasbx.abdm.gov.in/abha/api/v3",  "default", "ABHA service sandbox"),
    ("CLIENT_ID",        "(from .env)",                               "secret",  "ABDM_CLIENT_ID — set from Clinic Management Portal"),
    ("CLIENT_SECRET",    "(from .env)",                               "secret",  "ABDM_CLIENT_SECRET — never commit to git"),
    ("EMR_JWT",          "(auto-set by Login test script)",           "secret",  "JWT token from POST /api/emr/auth/login"),
    ("X_TOKEN",          "(auto-set by create-verify / login-verify)","secret",  "ABHA x-token returned after ABHA OTP flow"),
    ("ACCESS_TOKEN",     "(auto-set by debug/token)",                 "secret",  "ABDM gateway OAuth2 access token"),
    ("GATEWAY_TOKEN",    "(from ABDM sandbox portal)",                "secret",  "RS256 JWT sent by ABDM gateway on callbacks"),
    ("HIP_ID",           "(from .env ABDM_HIP_ID)",                  "default", "Your registered HIP ID in ABDM"),
    ("HIU_ID",           "(from .env)",                               "default", "Your registered HIU ID in ABDM"),
    ("PATIENT_ID",       "",                                          "default", "EMR patient UUID (emr_patients.id)"),
    ("ABHA_NUMBER",      "",                                          "default", "14-digit ABHA number e.g. 12-3456-7890-1234"),
    ("ABHA_ADDRESS",     "",                                          "default", "ABHA address e.g. patient@abdm"),
    ("CARE_CONTEXT_REF", "",                                          "default", "Care context reference number"),
    ("CONSENT_ID",       "",                                          "default", "Consent artefact UUID"),
    ("REQUEST_ID",       "(auto-set by test scripts)",                "default", "ABDM request UUID"),
    ("TRANSACTION_ID",   "(auto-set by test scripts)",                "default", "ABDM transaction UUID"),
    ("TXN_ID",           "(auto-set by OTP requests)",                "default", "ABHA service txnId"),
    ("LINK_REF_NUMBER",  "",                                          "default", "Link reference number from on-init callback"),
    ("AADHAAR_NUMBER",   "999941057058",                              "default", "ABDM sandbox test Aadhaar (pre-filled)"),
    ("MOBILE_NUMBER",    "",                                          "default", "10-digit mobile number"),
    ("APPOINTMENT_ID",   "",                                          "default", "EMR appointment UUID"),
]

for ri, (key, val, typ, desc) in enumerate(env_data, start=3):
    is_secret = typ == "secret"
    bg = "FFF2CC" if is_secret else ("F0F4FA" if ri % 2 == 0 else "FFFFFF")
    for ci, v in enumerate([key, val, typ, desc], start=1):
        c = ws_env.cell(row=ri, column=ci, value=v)
        c.fill = bfill(bg)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.font = Font(size=9, bold=(ci == 1), name="Courier New" if ci in (1,2) else "Calibri",
                      color=("C00000" if is_secret and ci == 3 else "000000"))
    ws_env.row_dimensions[ri].height = 18

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — ABDM Workflow Map
# ══════════════════════════════════════════════════════════════════════════════
ws_flow = wb.create_sheet("🔄 Workflow Map")
ws_flow.sheet_view.showGridLines = False

ws_flow.merge_cells("A1:F1")
ws_flow["A1"] = "ABDM HIP/HIU Workflow Sequence Map"
ws_flow["A1"].font = Font(bold=True, size=13, color=C_HEADER_FG)
ws_flow["A1"].fill = hfill(C_HEADER_BG)
ws_flow["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_flow.row_dimensions[1].height = 30

flow_cols = [
    ("Step #",     7),
    ("Milestone",  12),
    ("Actor",      18),
    ("Direction",  22),
    ("API Path",   55),
    ("Description",60),
]
for ci, (label, width) in enumerate(flow_cols, start=1):
    ws_flow.column_dimensions[get_column_letter(ci)].width = width
    c = ws_flow.cell(row=2, column=ci, value=label)
    c.font = Font(bold=True, size=10, color=C_HEADER_FG)
    c.fill = hfill("2E4057")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws_flow.row_dimensions[2].height = 22

flows = [
    # M1
    ("M1-1",  "M1", "Receptionist/EMR",  "EMR → ABHA Service",        "POST /api/emr/patients/:id/abha/create-otp",        "Generate Aadhaar OTP for ABHA creation"),
    ("M1-2",  "M1", "Receptionist/EMR",  "EMR → ABHA Service",        "POST /api/emr/patients/:id/abha/create-verify",     "Verify Aadhaar OTP, get txnId"),
    ("M1-3",  "M1", "Receptionist/EMR",  "EMR → ABHA Service",        "POST /api/emr/patients/:id/abha/mobile-otp",        "Send mobile OTP"),
    ("M1-4",  "M1", "Receptionist/EMR",  "EMR → ABHA Service",        "POST /api/emr/patients/:id/abha/mobile-verify",     "Verify mobile OTP — ABHA created"),
    ("M1-5",  "M1", "Receptionist/EMR",  "EMR → ABHA Service",        "POST /api/emr/patients/:id/abha/suggestions",       "Fetch ABHA address suggestions"),
    ("M1-6",  "M1", "Receptionist/EMR",  "EMR → ABHA Service",        "POST /api/emr/patients/:id/abha/set-address",       "Patient selects preferred ABHA address"),
    ("M1-7",  "M1", "Receptionist/EMR",  "EMR → ABHA Service",        "GET  /api/emr/patients/:id/abha/card",              "Download ABHA card PNG"),
    ("M1-8",  "M1", "Patient QR Scan",   "ABDM Gateway → HIP",        "POST /v3/hip/patient/share/profile",                "Patient scans HIP QR, shares profile"),
    ("M1-9",  "M1", "EMR (HIP)",         "HIP → HIECM",               "POST /v3/hip/patient/running-token/status",         "Verify patient token shown at reception"),
    # M2 Discovery
    ("M2-1",  "M2", "EMR (HIP)",         "EMR → DB",                  "POST /api/emr/patients/:id/care-contexts",          "Register care context in emr_care_contexts"),
    ("M2-2",  "M2", "EMR (HIU)",         "EMR → ABDM Gateway",        "POST /api/abdm/care-contexts/discover",             "Patient initiates discovery (async)"),
    ("M2-3",  "M2", "ABDM Gateway",      "Gateway → HIP",             "POST /v0.5/care-contexts/discover",                 "Gateway routes discovery to HIP"),
    ("M2-4",  "M2", "EMR (HIP)",         "HIP → Gateway (callback)",  "on-discover callback sent",                         "HIP returns matching care contexts"),
    ("M2-5",  "M2", "ABDM Gateway",      "Gateway → HIU",             "POST /v0.5/care-contexts/on-discover",              "Gateway delivers discovery result to HIU"),
    ("M2-6",  "M2", "EMR (HIU)",         "Poll",                      "GET  /api/abdm/care-contexts/discover/:requestId",  "HIU polls for discovery result"),
    # M2 Linking
    ("M2-7",  "M2", "EMR (HIU)",         "EMR → ABDM Gateway",        "POST /api/abdm/links/init",                         "Patient requests link init"),
    ("M2-8",  "M2", "ABDM Gateway",      "Gateway → HIP",             "POST /v0.5/links/link/init",                        "Gateway routes init to HIP"),
    ("M2-9",  "M2", "EMR (HIP)",         "HIP generates OTP (bcrypt)","hip_link_sessions table",                           "HIP creates secure OTP session"),
    ("M2-10", "M2", "ABDM Gateway",      "Gateway → HIU",             "POST /v0.5/links/link/on-init",                     "Gateway delivers linkRefNumber to HIU"),
    ("M2-11", "M2", "EMR (HIU)",         "EMR → ABDM Gateway",        "POST /api/abdm/links/confirm",                      "Patient submits OTP"),
    ("M2-12", "M2", "ABDM Gateway",      "Gateway → HIP",             "POST /v0.5/links/link/confirm",                     "Gateway routes OTP to HIP"),
    ("M2-13", "M2", "EMR (HIP)",         "HIP verifies bcrypt OTP",   "hip_link_sessions expiry + lockout",                "BLOCKER-1: expiry first, then lockout"),
    ("M2-14", "M2", "ABDM Gateway",      "Gateway → HIU",             "POST /v0.5/links/link/on-confirm",                  "Gateway delivers confirmed care contexts"),
    ("M2-15", "M2", "EMR (HIU/HIP)",     "HIP-initiated (v3)",        "POST /api/abdm/care-contexts/link",                 "EMR directly links via HIECM v3"),
    # M2 Consent
    ("M2-16", "M2", "EMR (HIU)",         "EMR → ABDM Gateway",        "POST /api/emr/consents",                            "HIU creates consent request"),
    ("M2-17", "M2", "ABDM Gateway",      "Gateway → HIU callback",    "POST /v0.5/consent-requests/on-init",               "Gateway maps ABDM consentRequest.id"),
    ("M2-18", "M2", "Patient (PHR App)", "Patient approves in PHR",   "(via PHR application)",                             "Patient grants/denies consent"),
    ("M2-19", "M2", "ABDM Gateway",      "Gateway → HIU",             "POST /v0.5/consents/hiu/notify",                    "Gateway notifies HIU: GRANTED/DENIED"),
    ("M2-20", "M2", "ABDM Gateway",      "Gateway → HIP",             "POST /v0.5/consents/hip/notify",                    "Gateway notifies HIP of consent status"),
    # M3 HI Exchange
    ("M3-1",  "M3", "EMR (HIU)",         "HIU → ABDM Gateway",        "(triggered by consent GRANTED)",                    "Auto-triggered: fetchHealthInfo() called"),
    ("M3-2",  "M3", "ABDM Gateway",      "Gateway → HIP",             "POST /v0.5/health-information/hip/request",         "Gateway requests health data from HIP"),
    ("M3-3",  "M3", "EMR (HIP)",         "HIP validates consent",     "hip_consent_artifacts + SEC-008",                   "Verify consent GRANTED, not expired"),
    ("M3-4",  "M3", "EMR (HIP)",         "HIP builds FHIR",           "buildFhirBundle() → fidelius-cli",                  "Build + encrypt with Curve25519/AES-256-GCM"),
    ("M3-5",  "M3", "EMR (HIP)",         "HIP → dataPushUrl",         "POST /v0.5/health-information/transfer",            "Push encrypted FHIR entries"),
    ("M3-6",  "M3", "ABDM Gateway",      "Gateway → HIU",             "POST /v0.5/health-information/transfer",            "Gateway forwards to HIU"),
    ("M3-7",  "M3", "EMR (HIU)",         "HIU decrypts FHIR",         "decryptHipEntry() → verify MD5",                    "Decrypt + verify checksum per §4.3.2"),
    ("M3-8",  "M3", "EMR (HIU)",         "Store",                     "health_records table",                              "Stored encrypted, decrypted on read"),
    ("M3-9",  "M3", "Clinician/EMR",     "EMR → DB",                  "GET  /api/abdm/health-records",                     "View decrypted health records"),
    ("M3-10", "M3", "❌ MISSING",         "HIU → Gateway (ACK)",       "/v0.5/health-information/notify",                   "Required ACK after receiving data — NOT IMPLEMENTED"),
]

milestone_bg = {"M1": C_M1, "M2": C_M2, "M3": C_M3}
for ri, (step, ms, actor, direction, path, desc) in enumerate(flows, start=3):
    bg = milestone_bg.get(ms, "FFF2CC") if "MISSING" not in step else C_DEBUG
    for ci, v in enumerate([step, ms, actor, direction, path, desc], start=1):
        c = ws_flow.cell(row=ri, column=ci, value=v)
        c.fill = bfill(bg)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.font = Font(size=9, bold=(ci == 1 or "MISSING" in str(v)),
                      color=("C00000" if "MISSING" in str(v) else "000000"),
                      name="Courier New" if ci == 5 else "Calibri")
    ws_flow.row_dimensions[ri].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Security & Gap Analysis
# ══════════════════════════════════════════════════════════════════════════════
ws_sec = wb.create_sheet("🔐 Security & Gaps")
ws_sec.sheet_view.showGridLines = False

ws_sec.merge_cells("A1:E1")
ws_sec["A1"] = "ABDM Security Audit & Gap Analysis"
ws_sec["A1"].font = Font(bold=True, size=13, color=C_HEADER_FG)
ws_sec["A1"].fill = hfill(C_HEADER_BG)
ws_sec["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sec.row_dimensions[1].height = 30

sec_cols = [("Severity", 12), ("Finding", 45), ("Location", 40), ("Risk / Impact", 55), ("Recommendation", 55)]
for ci, (label, width) in enumerate(sec_cols, start=1):
    ws_sec.column_dimensions[get_column_letter(ci)].width = width
    c = ws_sec.cell(row=2, column=ci, value=label)
    c.font = Font(bold=True, size=10, color=C_HEADER_FG)
    c.fill = hfill("2E4057")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws_sec.row_dimensions[2].height = 22

severity_colors = {
    "🔴 CRITICAL": "FFCCCC",
    "🟠 HIGH":     "FFE5CC",
    "🟡 MEDIUM":   "FFFACC",
    "✅ POSITIVE": "CCFFCC",
    "❌ MISSING":  "FFCCCC",
}

security_findings = [
    # CRITICAL
    ("🔴 CRITICAL", "Debug endpoints unauthenticated",
     "/api/abdm/debug/token\n/api/abdm/debug/bridge\n/api/abdm/debug/hip-sessions",
     "Any unauthenticated caller can obtain ABDM gateway access tokens and read patient OTP sessions",
     "Add emrAuth middleware to all 3 debug routes, or delete them before production"),

    ("🔴 CRITICAL", "Internal callback paths lack ABDM JWT verification",
     "/api/abdm/care-contexts/on-discover\n/api/abdm/links/link/on-init\n/api/abdm/links/link/on-confirm\n/api/abdm/consent/notify\n/api/abdm/health-info/push",
     "Anyone can POST fake ABDM callbacks — could inject fraudulent care context links or fake consent grants",
     "Add verifyAbdmCallback middleware to all 6 /api/abdm/* callback paths, or block them at nginx (allow only from ABDM IP range)"),

    # HIGH
    ("🟠 HIGH", "ABDM_SKIP_JWT_VERIFY must be false in production",
     "abdm.callback.auth.js\nprocess.env.ABDM_SKIP_JWT_VERIFY",
     "Bypasses all RS256 signature verification — allows forged ABDM gateway tokens",
     "Confirm ABDM_SKIP_JWT_VERIFY is not set or =false in production .env"),

    ("🟠 HIGH", "ABDM_DEV_SHOW_OTP=true logs OTP to console",
     "abdm.service.js\nprocess.env.ABDM_DEV_SHOW_OTP",
     "Console logs are captured by log aggregators — OTP leakage in aggregated logs",
     "Set ABDM_DEV_SHOW_OTP=false in production"),

    ("🟠 HIGH", "Rate limiter IP key may collide behind load balancer",
     "emr.routes.js otpLimiter\nkeyGenerator: (req) => req.ip",
     "If trust proxy not configured, all clients behind LB share one IP → rate limit bypassed by some, others DoS'd",
     "Configure app.set('trust proxy', 1) and verify req.ip resolves to real client IP"),

    # MEDIUM
    ("🟡 MEDIUM", "Health info rate limit is in-memory",
     "abdm.controller.js checkHealthInfoRateLimit()\n(in-memory Map)",
     "Counter resets on server restart; multi-instance deployments do not share state",
     "Use DB-backed rate limit (hip_rate_limits table already exists in hip.service.js) consistently"),

    ("🟡 MEDIUM", "CM_ID defaults to 'sbx' if not set",
     "abdm.service.js\nprocess.env.ABDM_CM_ID || 'sbx'",
     "If env var not set in production, sandbox CM ID used → consent artifacts routed incorrectly",
     "Set ABDM_CM_ID=abdm in production .env explicitly"),

    ("🟡 MEDIUM", "fidelius-cli path hardcoded default /opt/fidelius",
     "hip.service.js FIDELIUS_JAR",
     "If JAR missing at deploy time, all M3 health data push silently fails",
     "Add health-check at startup to verify fidelius JAR exists and is executable"),

    # MISSING APIS
    ("❌ MISSING", "POST /v0.5/health-information/notify not implemented",
     "server.js — no handler found",
     "ABDM spec requires HIU to ACK gateway after receiving /transfer data. Missing ACK may cause gateway to re-send or flag HI session as incomplete",
     "Implement handler that calls /v0.5/health-information/notify with sessionStatus=TRANSFERRED after storing records"),

    ("❌ MISSING", "WellnessRecord FHIR builder not implemented",
     "hip.service.js buildFhirBundle()",
     "HIP cannot push WellnessRecord HI type — consent for it will silently push no data",
     "Implement buildWellnessRecordBundle() in hip.service.js"),

    ("❌ MISSING", "HealthDocumentRecord FHIR builder not implemented",
     "hip.service.js buildFhirBundle()",
     "Cannot push scanned documents or PDFs as ABDM health records",
     "Implement buildHealthDocumentRecordBundle() using DocumentReference FHIR resource"),

    ("❌ MISSING", "ABHA HIU routes missing from abdm.routes.js",
     "/api/abdm/status, /api/abdm/profile, /api/abdm/logout, /api/abdm/card",
     "Endpoints referenced in ABDM HIU spec not registered — HIU cannot check linked ABHA status",
     "Add routes to abdm.routes.js pointing to abdmCtrl.getAbhaStatus, getAbhaProfile, getAbhaCard, logoutAbha"),

    # POSITIVES
    ("✅ POSITIVE", "OTP bcrypt-hashed, never stored plaintext", "hip.controller.js handleLinkInit", "SEC-022, R2-007 — bcrypt constant-time compare", "✅ Correct"),
    ("✅ POSITIVE", "OTP lockout after 3 attempts with expiry-first logic", "hip.controller.js handleLinkConfirm", "BLOCKER-1 fix applied", "✅ Correct"),
    ("✅ POSITIVE", "SSRF protection on dataPushUrl", "hip.service.js pushHealthData()", "R2-001 — allowlist *.abdm.gov.in only", "✅ Correct"),
    ("✅ POSITIVE", "PHI redacted in all logs", "Multiple locations", "R2-013, SEC-003, R2-008, R3-001", "✅ Correct"),
    ("✅ POSITIVE", "Fidelius-only encryption, no fallback", "hip.service.js", "R2-005 — aborts if JAR fails", "✅ Correct"),
    ("✅ POSITIVE", "MD5 checksum verification on received FHIR", "abdm.controller.js healthInfoPush()", "ABDM spec §4.3.2", "✅ Correct"),
    ("✅ POSITIVE", "UUID format validation on requestId/transactionId", "hip.controller.js", "BLOCKER-4 fix", "✅ Correct"),
    ("✅ POSITIVE", "SHA-256 hash of QR token stored (not plaintext)", "hip.controller.js handlePatientShareProfile()", "R3-007", "✅ Correct"),
    ("✅ POSITIVE", "Exponential backoff on gateway callbacks", "hip.service.js gwPostWithRetry()", "BLOCKER-5 fix", "✅ Correct"),
    ("✅ POSITIVE", "Consent validity verified before health data push", "hip.controller.js handleHealthInfoRequest()", "SEC-008, R3-009", "✅ Correct"),
]

for ri, (sev, finding, loc, risk, rec) in enumerate(security_findings, start=3):
    bg = severity_colors.get(sev, "FFFFFF")
    for ci, v in enumerate([sev, finding, loc, risk, rec], start=1):
        c = ws_sec.cell(row=ri, column=ci, value=v)
        c.fill = bfill(bg)
        c.border = thin_border()
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.font = Font(size=9, bold=(ci == 1),
                      color=("C00000" if "CRITICAL" in sev or "MISSING" in sev
                             else "007000" if "POSITIVE" in sev else "000000"))
    ws_sec.row_dimensions[ri].height = 52

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6 — Test Scripts Reference
# ══════════════════════════════════════════════════════════════════════════════
ws_test = wb.create_sheet("🧪 Test Scripts")
ws_test.sheet_view.showGridLines = False

ws_test.merge_cells("A1:D1")
ws_test["A1"] = "Postman Test Scripts — ABDM Collection"
ws_test["A1"].font = Font(bold=True, size=13, color=C_HEADER_FG)
ws_test["A1"].fill = hfill(C_HEADER_BG)
ws_test["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_test.row_dimensions[1].height = 30

test_cols = [("Request", 40), ("Test Assertion", 60), ("Variable Set", 30), ("Notes", 40)]
for ci, (label, width) in enumerate(test_cols, start=1):
    ws_test.column_dimensions[get_column_letter(ci)].width = width
    c = ws_test.cell(row=2, column=ci, value=label)
    c.font = Font(bold=True, size=10, color=C_HEADER_FG)
    c.fill = hfill("2E4057")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws_test.row_dimensions[2].height = 22

test_data = [
    ("Login (Get EMR JWT)",                    "Status 200, json.token exists",             "EMR_JWT ← json.token",            "Run first — all protected routes need this"),
    ("Create ABHA — 1. Gen Aadhaar OTP",       "Status 200, json.txnId is string",          "TXN_ID ← json.txnId",            "Rate-limited 3/10min"),
    ("Create ABHA — 2. Verify Aadhaar OTP",    "Status 200, json.txnId exists",             "TXN_ID, X_TOKEN",                "Saves updated txnId and xToken"),
    ("Create ABHA — 5. Get Suggestions",       "Status 200, suggestions array",             "(none)",                          ""),
    ("ABHA Login — 1. Request OTP",            "Status 200, json.txnId",                    "TXN_ID ← json.txnId",            ""),
    ("ABHA Login — 2. Verify OTP",             "Status 200, json.xToken",                   "X_TOKEN ← json.xToken",          ""),
    ("Add Patient via Aadhaar — 7. Finalize",  "Status 200 or 201",                         "PATIENT_ID ← json.patient.id",   "Creates EMR patient record"),
    ("Add Patient via ABHA — 2. Verify+Create","Status 200 or 201",                         "PATIENT_ID ← json.patient.id",   "Deduplicates by ABHA number"),
    ("Get Available Care Contexts",            "Status 200, array response",                 "(none)",                          ""),
    ("Add Care Context to Patient",            "Status 201, json.referenceNumber",           "CARE_CONTEXT_REF",               ""),
    ("Discover Care Contexts",                 "Status 200/202, json.requestId",             "REQUEST_ID ← json.requestId",    "Async — poll status next"),
    ("Poll Discovery Status",                  "Status 200, json.status is string",          "(none)",                          "Poll until status=COMPLETED"),
    ("Patient Link — 1. Init",                 "Status 200/202, json.requestId",             "REQUEST_ID ← json.requestId",    ""),
    ("Patient Link — 3. Confirm OTP",          "Status 200, json.requestId",                 "REQUEST_ID ← json.requestId",    ""),
    ("Create Consent Request",                 "Status 200/201, reqId/requestId",            "REQUEST_ID",                     ""),
    ("Pull Consent Health Records",            "Status 200/202, json.transactionId",         "TRANSACTION_ID",                 ""),
    ("Get Health Records (HIU View)",          "Status 200, records present",                "(none)",                          "Contains decrypted FHIR bundles"),
    ("HIP — Discover [v0.5]",                  "Status 202",                                 "(none)",                          "Simulates ABDM gateway calling your HIP"),
    ("HIP — Link Init [v0.5]",                 "Status 202",                                 "(none)",                          "HIP generates bcrypt OTP"),
    ("HIP — Link Confirm [v0.5]",              "Status 202",                                 "(none)",                          "HIP verifies OTP, validates care contexts"),
    ("HIP — Health Info Request [v0.5]",       "Status 202",                                 "(none)",                          "HIP builds FHIR, encrypts, pushes"),
    ("Test ABDM Gateway Token",                "Status 200, accessToken string",             "ACCESS_TOKEN",                   "⚠️ Debug only — remove in production"),
]

for ri, (req, assertion, var_set, notes) in enumerate(test_data, start=3):
    bg = C_ALT_ROW if ri % 2 == 0 else "FFFFFF"
    for ci, v in enumerate([req, assertion, var_set, notes], start=1):
        c = ws_test.cell(row=ri, column=ci, value=v)
        c.fill = bfill(bg)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.font = Font(size=9, name="Courier New" if ci in (2,3) else "Calibri")
    ws_test.row_dimensions[ri].height = 22

# ─── Save ─────────────────────────────────────────────────────────────────────
out_path = r"d:\Infer\ABDM_Postman_Collection.xlsx"
wb.save(out_path)
print("Saved: " + out_path)
print("API rows: " + str(len(rows)))
